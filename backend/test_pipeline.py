"""Quick sanity test against synthetic data covering every business rule,
including the ' ' (space) blank placeholder. Not a full test suite —
just enough to catch logic bugs before this goes to Claude Code."""

import io
import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from mtr_analysis import (
    Config, build_analysis_columns, reorder_to_final_layout,
    build_all_pivots, run_task1_trip_repush, run_task1_mapping,
    NEW_PLANT_DETENTION_SLAB, NEW_ZONE_REMARK,
    NEW_TRANSPORTER_REMARK, NEW_YARD_DETENTION_SLAB, NEW_AT_DEST_NAME,
    NEW_DEST_MATCH, NEW_DEST_DETENTION_SLAB, NEW_AT_SAP_LEAD_DIST, NEW_MATCH,
    NEW_AI_CHECK, NEW_SAP_AI, NEW_SAP_AI_REMARK, NEW_XSWIFT_PLANT_NAME_MATCH,
)

SP = " "  # the real blank placeholder

rows = [
    # row 0: fresh trip, everything blank via " " placeholder, plant detention = 0.
    # PlantName is a real value (not blank) here specifically so is_primary's
    # new NAME-based matching (first word "ADITYA") has something to match —
    # unlike the old Plant-Code-based matching, a blank name can never be
    # "primary" no matter what the code is.
    dict(TripID=1, Vehicle="V1", SAPPGI="P1", PGIDT="20 Jul 26 09:37", Transporter=SP, Zone=SP,
         YardIn=SP, YardOut=SP, YardDet=SP, PlantName="Aditya Loading", PlantCode="6910", PlantEntry="2026-07-20 09:37",
         PlantExit="2026-07-20 09:37", PlantDet="0", DestCode="CA12", Dest="AMBIKAPUR", DestEntry=SP,
         DestExit=SP, DestDet=SP, Stamp="Pending", SapLead="8", GPS=SP, AI=SP),
    # row 1: plant exit blank, entry present -> "vehicle still in plant"
    dict(TripID=2, Vehicle="V2", SAPPGI="P2", PGIDT="20 Jul 26 09:37", Transporter="ABC Transport", Zone="North B",
         YardIn="2026-07-20 03:47", YardOut="2026-07-20 03:57", YardDet="0:10", PlantName="Roorkee", PlantCode="6966",
         PlantEntry="2026-07-20 07:55", PlantExit=SP, PlantDet=SP, DestCode="VB46", Dest="BELDA", DestEntry=SP,
         DestExit=SP, DestDet=SP, Stamp="Pending", SapLead="21", GPS=SP, AI=SP),
    # row 2: stamp verified, both dest entry/exit present, detention 45 min -> slab "Above 30 min"
    dict(TripID=3, Vehicle="V3", SAPPGI="P3", PGIDT="20 Jul 26 09:37", Transporter="XYZ Transport", Zone="South A",
         YardIn="2026-07-20 01:00", YardOut="2026-07-20 01:00", YardDet="0", PlantName="Sidhi", PlantCode="6951",
         PlantEntry="2026-07-20 02:00", PlantExit="2026-07-20 03:00", PlantDet="15:00", DestCode="XX99", Dest="TESTDEST",
         DestEntry="2026-07-20 10:00", DestExit="2026-07-20 10:45", DestDet="0:45", Stamp="Stamp Verified",
         SapLead="100", GPS="95", AI="130"),  # SAP-AI = 100-130 = -30 -> "AI usages is high"
    # row 3: AI Repaired Distance literal "0" -> AI check "Not Available" but
    # SAP-AI Remark still bands normally off the real diff (50-0=50 -> low),
    # NOT "AI is blank" (real bug fix: literal 0 is a valid number, only
    # genuinely blank AI Repaired Distance gets "AI is blank" — see row 5).
    dict(TripID=4, Vehicle="V4", SAPPGI="P4", PGIDT="20 Jul 26 09:37", Transporter="Foo Carriers", Zone="East A",
         YardIn=SP, YardOut=SP, YardDet=SP, PlantName="Dalla", PlantCode="6633-3301",
         PlantEntry="2026-07-20 02:00", PlantExit="2026-07-20 03:00", PlantDet="10:00", DestCode="YY88", Dest="OTHERDEST",
         DestEntry="2026-07-20 10:00", DestExit="2026-07-20 10:05", DestDet="0:05", Stamp="Low Confidence",
         SapLead="50", GPS="48", AI="0"),
    # row 4: SAP-AI within -20..20 -> "0-20 "
    dict(TripID=5, Vehicle="V5", SAPPGI="P5", PGIDT="20 Jul 26 09:37", Transporter="Bar Logistics", Zone="Central A",
         YardIn=SP, YardOut=SP, YardDet=SP, PlantName="Vikram", PlantCode="6911",
         PlantEntry="2026-07-20 02:00", PlantExit="2026-07-20 03:00", PlantDet="10:00", DestCode="ZZ77", Dest="THIRDDEST",
         DestEntry="2026-07-20 10:00", DestExit="2026-07-20 10:05", DestDet="0:05", Stamp="Stamp Verified",
         SapLead="60", GPS="59", AI="55"),  # SAP-AI = 5 -> "0-20 "
    # row 5: AI Repaired Distance BLANK (not literal "0") -> "AI is blank" too
    # (real bug fix: this case was only catching literal 0, not blank)
    dict(TripID=6, Vehicle="V6", SAPPGI="P6", PGIDT="20 Jul 26 09:37", Transporter="Foo Carriers", Zone="East A",
         YardIn=SP, YardOut=SP, YardDet=SP, PlantName="Dalla", PlantCode="6633-3301",
         PlantEntry="2026-07-20 02:00", PlantExit="2026-07-20 03:00", PlantDet="10:00", DestCode="YY88", Dest="OTHERDEST",
         DestEntry="2026-07-20 10:00", DestExit="2026-07-20 10:05", DestDet="0:05", Stamp="Stamp Verified",
         SapLead="50", GPS="48", AI=SP),
    # row 6: Banswara Loading + blank Zone -> "Zone enable" (real bug fix:
    # this is a Banswara-only flag, not primary-plant-based)
    dict(TripID=7, Vehicle="V7", SAPPGI="P7", PGIDT="20 Jul 26 09:37", Transporter="Foo Carriers", Zone=SP,
         YardIn=SP, YardOut=SP, YardDet=SP, PlantName="Banswara Loading", PlantCode="9999",
         PlantEntry="2026-07-20 02:00", PlantExit="2026-07-20 03:00", PlantDet="10:00", DestCode="YY88", Dest="OTHERDEST",
         DestEntry="2026-07-20 10:00", DestExit="2026-07-20 10:05", DestDet="0:05", Stamp="Stamp Verified",
         SapLead="50", GPS="48", AI="45"),
    # row 7: Banswara Loading but Zone PRESENT -> Zone Remark stays blank
    dict(TripID=8, Vehicle="V8", SAPPGI="P8", PGIDT="20 Jul 26 09:37", Transporter="Foo Carriers", Zone="North A",
         YardIn=SP, YardOut=SP, YardDet=SP, PlantName="Banswara Loading", PlantCode="9999",
         PlantEntry="2026-07-20 02:00", PlantExit="2026-07-20 03:00", PlantDet="10:00", DestCode="YY88", Dest="OTHERDEST",
         DestEntry="2026-07-20 10:00", DestExit="2026-07-20 10:05", DestDet="0:05", Stamp="Stamp Verified",
         SapLead="50", GPS="48", AI="45"),
]

df_raw = pd.DataFrame(rows)
# Map synthetic columns onto real column names the pipeline expects
mtr = pd.DataFrame({
    "Trip ID": df_raw.TripID, "Vehicle No.": df_raw.Vehicle, "Vehicle Type": "1101",
    "SAP PGI No": df_raw.SAPPGI, "PGI Date & Time": df_raw.PGIDT, "SAP Order No": "1",
    "DI No": "1", "Transporter Name ": df_raw.Transporter, "Transporter Code": "1",
    "Zone": df_raw.Zone, "Yard IN": df_raw.YardIn, "Yard Out": df_raw.YardOut,
    "Yard detention": df_raw.YardDet, "Plant name": df_raw.PlantName, "Plant Code": df_raw.PlantCode,
    "Plant Entry": df_raw.PlantEntry, "Plant Exit": df_raw.PlantExit, "Plant Detention": df_raw.PlantDet,
    "Destination Code": df_raw.DestCode, "Destination": df_raw.Dest, "Customer Name": "cust",
    "Dest Entry Time": df_raw.DestEntry, "Dest Exit Time": df_raw.DestExit, "Dest Detention": df_raw.DestDet,
    "Destination Proximity End Time": SP, "Destination Ageing": SP, "Onward Duration": SP,
    "Customer Segment": "TRADE", "Compliance Status": SP, "Depot": SP, "Route Name": SP, "Halt": SP,
    "Onward Status": "Pending", "Stamp Status": df_raw.Stamp, "Reject Reason": SP,
    "Sap Lead Dist": df_raw.SapLead, "GPS Distance": df_raw.GPS, "AI Repaired Distance": df_raw.AI,
    "Geofence Hit/miss": "No", "Billing Status": "Pending",
    "Mother Geofence Start Time": SP, "Mother Geofence End Time": SP, "Mother Geofence Detention": SP,
})

city_code_to_destination = {"CA12": "AMBIKAPUR", "VB46": "BELDA"}  # ZZ77/YY88/XX99 intentionally missing -> #N/A
sap_pgi_to_lead_dist = {"P1": "8", "P2": "21", "P3": "100"}  # P4, P5 intentionally missing -> #N/A
# 2026-07-22: is_primary is now NAME-based (first word, case-insensitive) —
# these are first words of "AT Plant Name" tab entries, not Plant Codes.
# Matches rows 0/1/2/3/4's PlantName values (Aditya/Roorkee/Sidhi/Dalla/Vikram).
primary_plant_first_words = {"ADITYA", "ROORKEE", "SIDHI", "DALLA", "VIKRAM"}
# XSwift Plant Name reference map, keyed by Plant Code — row 0 (6910) has a
# matching registered name -> TRUE; row 1 (6966) has a deliberately WRONG
# registered name -> FALSE; other codes are absent -> "NA".
xswift_plant_name_map_test = {"6910": "Aditya Loading", "6966": "Some Other Name"}

cfg = Config(mtr_csv=Path("."), consignment_xlsx=Path("."), primary_plants_xlsx=Path("."))

result = build_analysis_columns(
    mtr, cfg, city_code_to_destination, sap_pgi_to_lead_dist,
    primary_plant_first_words, xswift_plant_name_map_test,
)
result = reorder_to_final_layout(result)

checks = [
    ("row0 Plant detention Slab == 'Loading not merged'", result.loc[0, NEW_PLANT_DETENTION_SLAB] == "Loading not merged"),
    ("row0 Transporter Remark == 'Not Available'", result.loc[0, NEW_TRANSPORTER_REMARK] == "Not Available"),
    ("row0 Yard Detention Slab is blank (Stamp=Pending, out of scope)", pd.isna(result.loc[0, NEW_YARD_DETENTION_SLAB])),
    ("row0 AT SAP lead distance == '8' (numeric match)", str(result.loc[0, NEW_AT_SAP_LEAD_DIST]) == "8"),
    ("row0 Match == 'TRUE'", result.loc[0, NEW_MATCH] == "TRUE"),
    ("row1 Plant detention Slab == 'vehicle still in plant'", result.loc[1, NEW_PLANT_DETENTION_SLAB] == "vehicle still in plant"),
    ("row1 Yard Detention Slab is blank (Stamp=Pending, out of scope)", pd.isna(result.loc[1, NEW_YARD_DETENTION_SLAB])),
    ("row2 Plant detention Slab == 'Above 30 min'", result.loc[2, NEW_PLANT_DETENTION_SLAB] == "Above 30 min"),
    ("row2 Yard Detention Slab == 'Not Available' (Stamp Verified, in/out present, detention 0)", result.loc[2, NEW_YARD_DETENTION_SLAB] == "Not Available"),
    ("row2 Destination detention slab == '30-45'? check band", result.loc[2, NEW_DEST_DETENTION_SLAB] == "Above 30 min"),
    ("row2 SAP-AI Remark == 'AI usages is high'", result.loc[2, NEW_SAP_AI_REMARK] == "AI usages is high"),
    ("row2 AT destination name == '#N/A' (ZZ.. not in lookup? no XX99)", result.loc[2, NEW_AT_DEST_NAME] == "#N/A"),
    ("row2 Dest. Match == 'NA'", result.loc[2, NEW_DEST_MATCH] == "NA"),
    ("row3 AI check == 'Not Available' (AI=0)", result.loc[3, NEW_AI_CHECK] == "Not Available"),
    ("row3 SAP-AI Remark == 'AI usages is low' (literal 0, not blank -> bands normally)", result.loc[3, NEW_SAP_AI_REMARK] == "AI usages is low"),
    ("row4 SAP-AI Remark == '0-20 '", result.loc[4, NEW_SAP_AI_REMARK] == "0-20 "),
    ("row4 AI check == 'Available'", result.loc[4, NEW_AI_CHECK] == "Available"),
    ("row0 XSwift Plant Name Match == 'TRUE' (registered name matches)", result.loc[0, NEW_XSWIFT_PLANT_NAME_MATCH] == "TRUE"),
    ("row1 XSwift Plant Name Match == 'FALSE' (registered name differs)", result.loc[1, NEW_XSWIFT_PLANT_NAME_MATCH] == "FALSE"),
    ("row2 XSwift Plant Name Match == 'NA' (code not in reference map)", result.loc[2, NEW_XSWIFT_PLANT_NAME_MATCH] == "NA"),
    ("row5 AI check == 'Not Available' (AI Repaired Distance BLANK, not literal 0)", result.loc[5, NEW_AI_CHECK] == "Not Available"),
    ("row5 SAP-AI Remark == 'AI is blank' (real bug fix: blank must count, not just literal 0)", result.loc[5, NEW_SAP_AI_REMARK] == "AI is blank"),
    ("row6 Zone Remark == 'Zone enable' (Banswara Loading + blank Zone)", result.loc[6, NEW_ZONE_REMARK] == "Zone enable"),
    ("row7 Zone Remark is blank (Banswara Loading but Zone present)", result.loc[7, NEW_ZONE_REMARK] == ""),
]

print(f"{'PASS' if all(c[1] for c in checks) else 'FAIL'} — {sum(c[1] for c in checks)}/{len(checks)} checks passed\n")
for desc, passed in checks:
    print(f"  [{'x' if passed else ' '}] {desc}")

pivots_by_sheet = build_all_pivots(result)
print(f"\nBuilt {sum(len(v) for v in pivots_by_sheet.values())} pivots across {len(pivots_by_sheet)} sheets:")
for sheet, pivots in pivots_by_sheet.items():
    for title, pivot_df in pivots:
        print(f"\n[{sheet}] {title}  (shape={pivot_df.shape})")
        print(pivot_df)

# --- Trip Repush test ---
# 3 synthetic consignment rows: one primary-plant + missing from MTR (should
# appear in Trip Repush), one primary-plant + present in MTR (should NOT
# appear), one non-primary-plant + missing from MTR (should NOT appear).
consignment = pd.DataFrame({
    "Company": ["Aditya Cement Works", "Aditya Cement Works", "Some Secondary Plant"],
    "SAP PGI No": ["P_MISSING_1", "P1", "P_MISSING_2"],  # P1 exists in mtr (row0 of synthetic mtr above)
    "Plant Code": ["6910", "6910", "9999"],  # 6910 is primary (Aditya), 9999 is not
})
repush = run_task1_trip_repush(consignment, mtr, primary_plant_first_words)
repush_checks = [
    ("Trip Repush includes primary+missing (P_MISSING_1)", "P_MISSING_1" in set(repush["SAP PGI No"])),
    ("Trip Repush excludes primary+present-in-MTR (P1)", "P1" not in set(repush["SAP PGI No"])),
    ("Trip Repush excludes non-primary+missing (P_MISSING_2)", "P_MISSING_2" not in set(repush["SAP PGI No"])),
    ("Trip Repush row count == 1", len(repush) == 1),
]
print(f"\n--- Trip Repush checks ---")
for desc, passed in repush_checks:
    print(f"  [{'x' if passed else ' '}] {desc}")
checks.extend(repush_checks)

print(f"\n{'=' * 50}")
print(f"FINAL: {'PASS' if all(c[1] for c in checks) else 'FAIL'} — {sum(c[1] for c in checks)}/{len(checks)} total checks passed")

# --- Mapping issue test ---
# Covers the real bug found: a vehicle with a BLANK Vehicle Reg Plant Name
# in XSwift must still be treated as primary (XSwift-side plant filter
# removed entirely — see run_task1_mapping docstring).
import tempfile

with tempfile.TemporaryDirectory() as tmpdir:
    tmpdir = Path(tmpdir)

    # XSwift "Trip Dashboard" sheet: 2 banner rows + header + data
    xswift_data = pd.DataFrame({
        "Vehicle No": ["V_ONLY_XSWIFT_OFFLINE", "V_ONLY_XSWIFT_ONLINE", "V_BOTH", "V_BLANK_PLANT_OFFLINE"],
        "Vehicle Status": ["Offline", "Online", "Idle", "Offline"],
        "Vehicle Reg Plant Name": ["ADITYA CEMENT WORKS_UTCL(P)", "ADITYA CEMENT WORKS_UTCL(P)", "ADITYA CEMENT WORKS_UTCL(P)", ""],
    })
    xswift_path = tmpdir / "xswift.xlsx"
    with pd.ExcelWriter(xswift_path) as w:
        # write 2 blank banner rows above the real header, matching real file structure
        pd.DataFrame([[None]]).to_excel(w, sheet_name="Trip Dashboard", header=False, index=False, startrow=0)
        pd.DataFrame([[None]]).to_excel(w, sheet_name="Trip Dashboard", header=False, index=False, startrow=1)
        xswift_data.to_excel(w, sheet_name="Trip Dashboard", index=False, startrow=2)

    # AT "dashboard" sheet
    at_data = pd.DataFrame({
        "Company Name": ["Aditya Cement Works", "Aditya Cement Works", "Some Non Primary Co", "Aditya Raw Material"],
        "Share": ["Veh Share", "Veh Share", "Veh Share", "Veh Share"],
        "Vehicle": ["V_BOTH", "V_ONLY_AT", "V_NONPRIMARY_ONLY_AT", "V_ONLY_AT_NAME_VARIANT"],
        "Status": ["Idle", "Idle", "Idle", "Idle"],
    })
    at_path = tmpdir / "at.xlsx"
    at_data.to_excel(at_path, sheet_name="dashboard", index=False)

    cfg_mapping = Config(
        mtr_csv=Path("."), consignment_xlsx=Path("."), primary_plants_xlsx=Path("."),
        xswift_live_dashboard_xlsx=xswift_path, at_live_dashboard_xlsx=at_path,
    )
    mapping_result = run_task1_mapping(cfg_mapping, primary_plant_first_words={"ADITYA"})

    not_in_at = set(mapping_result["Not in AT"]["Vehicle No"])
    not_in_swift = set(mapping_result["Not in Swift"]["Vehicle"])

    mapping_checks = [
        ("Not in AT includes offline-only-in-xswift vehicle", "V_ONLY_XSWIFT_OFFLINE" in not_in_at),
        ("Not in AT EXCLUDES online-only-in-xswift vehicle (status filter)", "V_ONLY_XSWIFT_ONLINE" not in not_in_at),
        ("Not in AT excludes vehicle present on both", "V_BOTH" not in not_in_at),
        ("Not in AT includes BLANK-plant-name vehicle (the real bug fix)", "V_BLANK_PLANT_OFFLINE" in not_in_at),
        ("Not in Swift includes AT-only primary-plant vehicle", "V_ONLY_AT" in not_in_swift),
        ("Not in Swift EXCLUDES non-primary-plant AT vehicle", "V_NONPRIMARY_ONLY_AT" not in not_in_swift),
        ("Not in Swift excludes vehicle present on both", "V_BOTH" not in not_in_swift),
        ("Not in Swift includes AT company-name-variant vehicle (first-word match fix)", "V_ONLY_AT_NAME_VARIANT" in not_in_swift),
    ]
    print(f"\n--- Mapping issue checks ---")
    for desc, passed in mapping_checks:
        print(f"  [{'x' if passed else ' '}] {desc}")
    checks.extend(mapping_checks)

print(f"\n{'=' * 50}")
print(f"GRAND FINAL: {'PASS' if all(c[1] for c in checks) else 'FAIL'} — {sum(c[1] for c in checks)}/{len(checks)} total checks passed")

# --- Regression test: xlsx write must not silently drop columns ---
# Guards against a real, previously-shipped bug: xlsxwriter's
# `constant_memory: True` option silently wrote only the FIRST column of
# a DataFrame and left every other column blank, with no error raised.
# This test fails loudly if that regression is ever reintroduced.
import tempfile
from mtr_analysis import write_xlsx
from openpyxl import load_workbook

with tempfile.TemporaryDirectory() as tmpdir:
    test_df = pd.DataFrame({
        "Trip ID": [1, 2, 3],
        "Vehicle No.": ["V1", "V2", "V3"],
        "Plant name": ["Plant A", "Plant B", "Plant C"],
        NEW_TRANSPORTER_REMARK: ["Available", "Not Available", "Available"],
    })
    test_path = Path(tmpdir) / "regression_test.xlsx"
    write_xlsx("main", test_df, {}, test_path)

    wb = load_workbook(test_path, read_only=True, data_only=True)
    written_rows = list(wb["main"].iter_rows(max_row=4, values_only=True))
    wb.close()

    # Check formatting directly from the workbook's styles (non-read-only load)
    from openpyxl import load_workbook as load_wb_full
    wb_full = load_wb_full(test_path)
    ws_full = wb_full["main"]
    # Find the "Transporter Remark" column (col D, index 4, 1-based) and
    # check its header cell has a yellow fill; "Trip ID" (col A) should not.
    transporter_remark_col_idx = list(test_df.columns).index(NEW_TRANSPORTER_REMARK) + 1
    trip_id_header_fill = ws_full.cell(row=1, column=1).fill.fgColor.rgb
    transporter_remark_header_fill = ws_full.cell(row=1, column=transporter_remark_col_idx).fill.fgColor.rgb
    wb_full.close()

    regression_checks = [
        ("xlsx write: header row intact", written_rows[0] == ("Trip ID", "Vehicle No.", "Plant name", NEW_TRANSPORTER_REMARK)),
        ("xlsx write: row 1 fully populated (not just first column)", written_rows[1] == (1, "V1", "Plant A", "Available")),
        ("xlsx write: row 2 fully populated (not just first column)", written_rows[2] == (2, "V2", "Plant B", "Not Available")),
        ("xlsx write: row 3 fully populated (not just first column)", written_rows[3] == (3, "V3", "Plant C", "Available")),
        ("xlsx formatting: Trip ID header NOT yellow", trip_id_header_fill != "FFFFFF00"),
        ("xlsx formatting: Transporter Remark header IS yellow", transporter_remark_header_fill == "FFFFFF00"),
    ]
    print(f"\n--- XLSX corruption regression checks ---")
    for desc, passed in regression_checks:
        print(f"  [{'x' if passed else ' '}] {desc}")
    checks.extend(regression_checks)

# --- Column validation checks ---
# Confirms validate_inputs() correctly reports missing vs. extra columns
# per-file, and only raises when something REQUIRED is actually missing.
from mtr_analysis import validate_inputs, ColumnValidationError, REQUIRED_MTR_CSV_COLUMNS, REQUIRED_CONSIGNMENT_COLUMNS

_bad_cols = [c for c in REQUIRED_MTR_CSV_COLUMNS if c != "Stamp Status"] + ["Weird New Field"]
_bad_mtr = pd.DataFrame([["x"] * len(_bad_cols)], columns=_bad_cols).to_csv(index=False).encode()

_ok_cons_cols = REQUIRED_CONSIGNMENT_COLUMNS + ["Vehicle"]
_ok_cons_df = pd.DataFrame([["y"] * len(_ok_cons_cols)], columns=_ok_cons_cols)
_cons_buf = io.BytesIO()
with pd.ExcelWriter(_cons_buf, engine="xlsxwriter") as _w:
    _ok_cons_df.to_excel(_w, sheet_name="Consignment Report", index=False)

validation_checks = []
try:
    validate_inputs(_bad_mtr, _cons_buf.getvalue())
    validation_checks.append(("validate_inputs raises on missing required column", False))
except ColumnValidationError as e:
    validation_checks.append(("validate_inputs raises on missing required column", True))
    validation_checks.append(("error report flags 'Stamp Status' as missing", "Stamp Status" in e.report["Raw MTR CSV"]["missing"]))
    validation_checks.append(("error report flags 'Weird New Field' as extra", "Weird New Field" in e.report["Raw MTR CSV"]["extra"]))
    validation_checks.append(("error report does NOT flag consignment as missing anything (extra-only is fine)", e.report["AT Consignment Report"]["missing"] == []))
    validation_checks.append(("error message mentions the file label", "[Raw MTR CSV]" in str(e)))

_ok_mtr_df = pd.DataFrame([["x"] * len(REQUIRED_MTR_CSV_COLUMNS)], columns=REQUIRED_MTR_CSV_COLUMNS)
_ok_mtr = _ok_mtr_df.to_csv(index=False).encode()
try:
    report = validate_inputs(_ok_mtr, _cons_buf.getvalue())
    validation_checks.append(("validate_inputs passes (no exception) when all required columns present", True))
except ColumnValidationError:
    validation_checks.append(("validate_inputs passes (no exception) when all required columns present", False))

print(f"\n--- Column validation checks ---")
for desc, passed in validation_checks:
    print(f"  [{'x' if passed else ' '}] {desc}")
checks.extend(validation_checks)

print(f"\n{'=' * 50}")
print(f"ABSOLUTE FINAL: {'PASS' if all(c[1] for c in checks) else 'FAIL'} — {sum(c[1] for c in checks)}/{len(checks)} total checks passed")
